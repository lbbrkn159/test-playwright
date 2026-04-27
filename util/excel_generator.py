import os
from robot import run
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.drawing.image import Image

def generate_excel(test_results):
    script_dir = os.path.dirname(__file__)
    excel_path = os.path.join(script_dir, "TestCases.xlsx")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # Add header row
    ws.append([ "Module", "Menu", "Function", "Test Case ID", "Alternative Test Case",
                "Preconditions", "Dependencies", "Ref Other Module (IN)", "Ref Other Module (OUT)",
                "Test Scenario", "Test Description", "Test Data", "Test Step", "Tester / User Role",
                "Test Script ID", "Test Script Name", "Alternate Flow", "Expected Result", "Actual Result",
                "Test Environment", "Browser", "Priority", "Test Status", "Execution Date", "Setting Result"])

    # Set column widths
    column_widths = [20, 20, 20, 20, 20, 30, 20, 20, 20, 30, 50, 50, 50, 20, 20, 30, 20, 30, 30, 20, 20, 10, 10, 20, 20]
    for col_num, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + col_num)].width = width
    
    for test_data in test_results:
        row = [test_data[key].strip() for key in test_data]
        ws.append(row)
        
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Highlight the test status cell
        test_status_cell = ws.cell(row=ws.max_row, column=ws.max_column - 2)
        if test_data["test_status"] == "PASS":
            test_status_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        elif test_data["test_status"] == "FAIL":
            screenshot_path = test_data.get("screenshot_path")
            if screenshot_path and os.path.exists(screenshot_path):
                img = Image(screenshot_path)
                img.width = 200
                img.height = 150
                ws.add_image(img, f"C{ws.max_row}")
            test_status_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    wb.save(excel_path)
    print(f"Created file {excel_path} successfully!")
